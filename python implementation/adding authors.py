from openpyxl import load_workbook
wb = load_workbook(filename = 'updatedpapersISME.xlsx')
sheet_ranges = wb.active


f = open("author table.txt", "w")


class paperClass(object):
    def __init__(self, number, authName, topic, sno):
        self.number = number
        self.authName = authName
        self.topic = topic
        self.sno =  sno



paperObjs = []

k=1
paperObjs.append(paperClass(10, "Sanjeev Kumar Gupta, R.C. Mehta",
        "Empirical Formulation of Sequent Depth Ratio and Relative Height of Hydraulic Jump in Sloping Prismatic Channel",
        k))

paperObjs.append(paperClass(52, "Vaibhav Dhar Dwivedi, Pankaj Wahi",
        "Influence of strut top mount bushing flexibility on the performance of suspension system",
        k))

k=1
for j in range(1, 16, 5):
    for i in range(13, 19):
        paperObjs.append(paperClass(sheet_ranges.cell(row=i, column=j).value, sheet_ranges.cell(row=i, column=j+2).value,
        sheet_ranges.cell(row=i, column=j+3).value,
        k))
        k+=1

for j in range(1, 16, 5):
    for i in range(20, 27):
        paperObjs.append(paperClass(sheet_ranges.cell(row=i, column=j).value, sheet_ranges.cell(row=i, column=j+2).value,
        sheet_ranges.cell(row=i, column=j+3).value,
        k))
        k+=1

for j in range(1, 16, 5):
    for i in range(34, 41):
        paperObjs.append(paperClass(sheet_ranges.cell(row=i, column=j).value, sheet_ranges.cell(row=i, column=j+2).value,
        sheet_ranges.cell(row=i, column=j+3).value,
        k))
        k+=1

for j in range(1, 11, 5):
    for i in range(48, 66):
        paperObjs.append(paperClass(sheet_ranges.cell(row=i, column=j).value, sheet_ranges.cell(row=i, column=j+2).value,
        sheet_ranges.cell(row=i, column=j+3).value,
        k))
        k+=1

for j in range(11, 12, 5):
    for i in range(48, 62):
        paperObjs.append(paperClass(sheet_ranges.cell(row=i, column=j).value, sheet_ranges.cell(row=i, column=j+2).value,
        sheet_ranges.cell(row=i, column=j+3).value,
        k))
        k+=1

paperObjs = filter(lambda x: x.number!=None, paperObjs)

paperObjs = sorted(paperObjs, key=lambda obj: obj.number)

for obj in paperObjs:
    if not obj.authName:
        obj.authName = "N/A"
    obj.authName = obj.authName.replace("[&]", "&amp;")
    print(obj.number, " " + obj.authName + " " + obj.topic + "\n")

i=1
for obj in paperObjs:
    f.write("\n<tr>" + "\n\t" + "<td>" + str(i) + "</td>\n\t" + "<td>" + str(obj.number) + "</td>\n\t" + "<td>" + obj.authName + "</td>\n\t" + "<td><a target='_blank' href='pdfs/ISME2018_paper_" + str(obj.number) + ".pdf'>" + obj.topic + "</a></td>\n</tr>\n")
    i+=1
