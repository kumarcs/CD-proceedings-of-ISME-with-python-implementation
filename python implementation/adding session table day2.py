from openpyxl import load_workbook
wb = load_workbook(filename = 'updatedpapersISME.xlsx')
sheet_ranges = wb.active


f = open("adding session table day2.txt", "w")


for i in range(1, 6):
    print(sheet_ranges.cell(row=14, column=i).value)

class paperClass(object):
    def __init__(self, number, authName, topic):
        self.number = number
        self.authName = authName
        self.topic = topic



paperObjs = []

k=1
for j in range(1, 16, 5):
    for i in range(34, 41):
        paperObjs.append(paperClass(sheet_ranges.cell(row=i, column=j).value, sheet_ranges.cell(row=i, column=j+2).value,
        sheet_ranges.cell(row=i, column=j+3).value))

for i in range(0, len(paperObjs)):
    print(paperObjs[i].authName)

k=1
print(len(paperObjs))
for i in range(0, len(paperObjs), 7):
    print(str(i) + "\n");
    f.write("<tr>\n<td colspan='9' bgcolor='#E9EACE'><div align='center' class='track'>Track " + str(k) + " Room:LT-1/12:15 PM-01:15 PM</div></td>\n</tr>\n")
    f.write("<tr> \n <th bgcolor='#E9EACE'><div align='center'>Manuscript Id</div></th> <th colspan='3' bgcolor='#E9EACE'><div align='center'>Author's Name</div></th><th colspan='5' bgcolor='#E9EACE'><div align='center'>Topic</div></th>\n</tr>")
    for j in range(i, i+7):
        print(str(j) + " ");
        f.write("<tr>\n<td bgcolor='#E9EACE'><div align='center'>" + str(paperObjs[j].number) + "</div></td>\n<td colspan='3' bgcolor='#E9EACE'><div align='center'>" + paperObjs[j].authName + "</div></td> \n <td colspan='5' bgcolor='#E9EACE'><div align='center'><a target='_blank' href='pdfs/ISME2018_paper_" + str(paperObjs[j].number) + ".pdf'>" + paperObjs[j].topic + "</a></div></td>\n</tr>")
    print("\n")
    k+=1;
